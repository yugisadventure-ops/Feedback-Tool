#!/usr/bin/env python3
"""
Prepare a feedback Excel template for the S2 Launch Feedback Tool.

Extracts embedded images to assets/images/ and writes a normalized template
with image paths the web app can load.

Usage:
  python3 scripts/prepare-template.py path/to/your-feedback-form.xlsx
"""

from __future__ import annotations

import re
import shutil
import sys
import zipfile
from pathlib import Path

try:
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
except ImportError:
    print("Install dependencies: pip install openpyxl pillow")
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
ASSETS_DIR = ROOT / "assets" / "images"
OUTPUT_TEMPLATE = ROOT / "templates" / "feedback-form.xlsx"


def slugify(text: str) -> str:
    s = re.sub(r"[^a-zA-Z0-9]+", "-", text.strip().lower()).strip("-")
    return s[:60] or "image"


def extract_images_from_xlsx(src: Path) -> dict[tuple[int, int], str]:
    """Return map of (row, col) -> saved image path relative to repo root."""
    ASSETS_DIR.mkdir(parents=True, exist_ok=True)
    mapping: dict[tuple[int, int], str] = {}

    with zipfile.ZipFile(src) as zf:
        media = [n for n in zf.namelist() if n.startswith("xl/media/")]
        for i, media_path in enumerate(sorted(media), 1):
            ext = Path(media_path).suffix or ".png"
            out_name = f"q-image-{i}{ext}"
            out_path = ASSETS_DIR / out_name
            out_path.write_bytes(zf.read(media_path))
            mapping[(i, 0)] = f"assets/images/{out_name}"

    return mapping


def normalize_workbook(src: Path, dest: Path) -> None:
    wb = openpyxl.load_workbook(src)
    sheet = wb.active

    headers = [str(c.value or "").strip().lower() for c in sheet[1]]
    col_map = {h: i for i, h in enumerate(headers) if h}

    # Ensure standard columns exist
    required = ["id", "section", "type", "question", "options", "image", "required", "min", "max", "placeholder"]
    if not col_map:
        for i, h in enumerate(required, 1):
            sheet.cell(row=1, column=i, value=h)
        col_map = {h: i - 1 for i, h in enumerate(required, 0)}

    image_col = col_map.get("image", 5)

    # openpyxl anchor-based image extraction per sheet
    img_count = 0
    for img in getattr(sheet, "_images", []):
        anchor = img.anchor
        row = anchor._from.row + 1  # 1-based
        col = anchor._from.col
        img_count += 1
        ext = ".png"
        if hasattr(img, "path") and img.path:
            ext = Path(str(img.path)).suffix or ext
        out_name = f"row-{row}-col-{col}{ext}"
        out_path = ASSETS_DIR / out_name
        out_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(img.path, out_path) if hasattr(img, "path") and img.path else out_path.write_bytes(img._data())
        sheet.cell(row=row, column=image_col + 1, value=f"assets/images/{out_name}")

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    print(f"Saved normalized template: {dest}")
    print(f"Images extracted: {img_count}")


def main() -> None:
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    src = Path(sys.argv[1]).resolve()
    if not src.exists():
        print(f"File not found: {src}")
        sys.exit(1)

    normalize_workbook(src, OUTPUT_TEMPLATE)
    print("Done. The app will auto-load templates/feedback-form.xlsx on next refresh.")


if __name__ == "__main__":
    main()
